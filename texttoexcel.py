#!/usr/bin/env python3
"""
Convert pipe-separated employee data to Excel spreadsheet
Usage: python table_to_excel.py input.txt output.xlsx
"""

import sys
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def parse_table_data(text_content):
    """Parse both space-separated and pipe-separated table data"""
    lines = text_content.split('\n')
    
    data = []
    headers = ['First Name', 'Last Name', 'SS#', 'Date Hired']
    
    # Track if we've seen the data section start
    in_data_section = False
    
    for line in lines:
        # Skip empty lines
        if not line.strip():
            continue
        
        # Skip header separator lines (lines that are ONLY separators)
        if '---' in line and not re.search(r'\d{2}/\d{2}/\d{4}', line):
            # Only skip if it's a pure separator line (no dates in it)
            continue
        
        # Check if we're at the start of the data section
        if 'First Name' in line and 'Last Name' in line and 'SS#' in line:
            in_data_section = True
            continue
        
        # Method 1: Check for pipe-separated data
        # But first, check if this line starts with space-separated data followed by pipes
        # We want to prioritize space-separated parsing for the first part of the line
        if '|' in line and not re.match(r'^\s*\|', line):
            # Line contains pipes but doesn't start with a pipe
            # This might be space-separated data followed by pipe-separated data on the same line
            # Split the line at the first pipe cluster
            before_pipes = re.split(r'\s*\|', line)[0].strip()
            
            # Try to parse the part before pipes as space-separated
            if before_pipes and in_data_section:
                parts = before_pipes.split()
                if len(parts) >= 4:
                    potential_date = parts[-1]
                    potential_ssn = parts[-2]
                    ssn_match = re.match(r'^(xxx-xx-\S+|[\d\*]{3}-[\d\*]{2}-[\d\*]{4})$', potential_ssn)
                    date_match = re.match(r'^\d{2}/\d{2}/\d{4}$', potential_date)
                    
                    if ssn_match and date_match:
                        name_parts = parts[:-2]
                        if len(name_parts) >= 2:
                            last_name = name_parts[-1]
                            first_name = ' '.join(name_parts[:-1])
                            data.append([first_name, last_name, potential_ssn, potential_date])
            
            # Now process any pipe-separated data on the same line
            # Get everything after the first space-separated data
            pipe_section = line[line.index('|'):] if '|' in line else ''
            pipe_parts = [part.strip() for part in pipe_section.split('|')]
            pipe_parts = [p for p in pipe_parts if p and '---' not in p and not re.match(r'^\s*$', p)]
            
            # Look for sequences of 4 parts in the pipe_parts
            i = 0
            while i + 3 < len(pipe_parts):
                candidate = pipe_parts[i:i+4]
                # Validate this looks like a data row
                if (len(candidate) == 4 and 
                    candidate[0] and 
                    candidate[0] not in ['First Name', ''] and
                    re.match(r'^(xxx-xx-\S+|[\d\*]{3}-[\d\*]{2}-[\d\*]{4})$', candidate[2]) and
                    re.match(r'^\d{2}/\d{2}/\d{4}$', candidate[3])):
                    data.append(candidate)
                    i += 4
                else:
                    i += 1
                    
        elif '|' in line and re.match(r'^\s*\|', line):
            # Line starts with a pipe - pure pipe-separated data
            parts = [part.strip() for part in line.split('|')]
            parts = [p for p in parts if p]
            
            if len(parts) == 4:
                if parts[0] and parts[0] not in ['First Name', '']:
                    data.append(parts)
        
        # Method 2: Pure space-separated data (no pipes in line)
        elif in_data_section:
            parts = line.strip().split()
            
            if len(parts) >= 4:
                potential_date = parts[-1]
                potential_ssn = parts[-2]
                
                ssn_match = re.match(r'^(xxx-xx-\S+|[\d\*]{3}-[\d\*]{2}-[\d\*]{4})$', potential_ssn)
                date_match = re.match(r'^\d{2}/\d{2}/\d{4}$', potential_date)
                
                if ssn_match and date_match:
                    name_parts = parts[:-2]
                    
                    if len(name_parts) >= 2:
                        last_name = name_parts[-1]
                        first_name = ' '.join(name_parts[:-1])
                        data.append([first_name, last_name, potential_ssn, potential_date])
                    elif len(name_parts) == 1:
                        data.append(['', name_parts[0], potential_ssn, potential_date])
    
    return headers, data

def create_excel(headers, data, output_file):
    """Create an Excel file with the data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Data"
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    for row_idx, record in enumerate(data, start=2):
        for col_idx, value in enumerate(record, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            # Add borders
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Adjust column widths
    column_widths = [18, 18, 18, 15]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    # Save the workbook
    wb.save(output_file)
    print(f"Excel file created successfully: {output_file}")
    print(f"Total records processed: {len(data)}")

def main():
    if len(sys.argv) < 3:
        print("Usage: python table_to_excel.py input.txt output.xlsx [-v for verbose]")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    verbose = len(sys.argv) > 3 and sys.argv[3] == '-v'
    
    # Read the input file
    try:
        with open(input_file, 'r', encoding='utf-8') as f:
            text_content = f.read()
    except FileNotFoundError:
        print(f"Error: Input file '{input_file}' not found")
        sys.exit(1)
    
    # Parse and create Excel
    headers, data = parse_table_data(text_content)
    
    if not data:
        print("Warning: No data found in the input file")
        sys.exit(1)
    
    create_excel(headers, data, output_file)

if __name__ == "__main__":
    main()